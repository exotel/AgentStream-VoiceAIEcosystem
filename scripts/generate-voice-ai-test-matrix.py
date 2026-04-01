#!/usr/bin/env python3
"""
Generate docs/testing/exotel-voice-ai-test-matrix.xlsx.

Run from repo root:
  .venv-xlsx/bin/python scripts/generate-voice-ai-test-matrix.py
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "testing" / "exotel-voice-ai-test-matrix.xlsx"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize_columns(ws, max_width: int = 55) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col[:200])
        ws.column_dimensions[letter].width = min(max_width, max(12, length + 2))


PROVIDERS = [
    {
        "name": "ElevenLabs",
        "article": "docs/support/exotel-elevenlabs-sip-trunk.md",
        "pattern": "SIP digest + dashboard import",
        "dashboard": "https://elevenlabs.io/app/agents",
        "docs": "https://elevenlabs.io/docs/agents-platform/phone-numbers/sip-trunking",
        "notes": "TCP/TLS to Exotel edge; optional static-IP ACL",
    },
    {
        "name": "LiveKit Cloud",
        "article": "docs/support/exotel-livekit-sip-trunk.md",
        "pattern": "SIP telephony + inbound trunk + dispatch",
        "dashboard": "https://cloud.livekit.io/",
        "docs": "https://docs.livekit.io/telephony/start/sip-trunk-setup/",
        "notes": "Agent must join same room as SIP participant",
    },
    {
        "name": "Retell AI",
        "article": "docs/support/exotel-retell-sip-trunk.md",
        "pattern": "Elastic SIP / custom telephony",
        "dashboard": "https://dashboard.retellai.com/",
        "docs": "https://docs.retellai.com/deploy/custom-telephony",
        "notes": "sip.retellai.com destination; digest align",
    },
    {
        "name": "Bolna",
        "article": "docs/support/exotel-bolna-sip-trunk.md",
        "pattern": "BYOT SIP or Exotel dashboard connector",
        "dashboard": "https://www.bolna.ai/byot-setup",
        "docs": "https://www.bolna.ai/docs/sip-trunking/introduction",
        "notes": "SIP Beta; no SRTP; confirm inbound host in current Bolna docs",
    },
    {
        "name": "Pipecat + Daily",
        "article": "docs/support/exotel-pipecat-sip-trunk.md",
        "pattern": "Webhook + Daily room sip_uri bridge",
        "dashboard": "https://dashboard.daily.co/",
        "docs": "https://docs.pipecat.ai/guides/telephony/twilio-daily-sip",
        "notes": "Not a fixed single SIP host; engineering-heavy vs static destination-uris",
    },
    {
        "name": "Ultravox",
        "article": "docs/support/exotel-ultravox-sip-trunk.md",
        "pattern": "SIP (allowlist/registration) or native exotel medium",
        "dashboard": "https://app.ultravox.ai/",
        "docs": "https://docs.ultravox.ai/telephony/sip",
        "notes": "Alternate: Voice Streaming / telephony-platforms — no vSIP",
    },
    {
        "name": "Vapi",
        "article": "docs/support/exotel-vapi-sip-trunk.md",
        "pattern": "byo-sip-trunk + byo-phone-number",
        "dashboard": "https://dashboard.vapi.ai/",
        "docs": "https://docs.vapi.ai/advanced/sip/sip-trunk",
        "notes": "Gateway IPv4; whitelist both Vapi SBC /32 IPs on Exotel",
    },
    {
        "name": "Smallest AI (Atoms)",
        "article": "docs/support/exotel-smallest-ai-sip-trunk.md",
        "pattern": "Import SIP (termination + origination URL)",
        "dashboard": "https://app.smallest.ai/",
        "docs": "https://atoms-docs.smallest.ai/platform/deployment/phone-numbers.md",
        "notes": "Copy SIP Origination URL from UI into Exotel destination-uris",
    },
    {
        "name": "Vocallabs",
        "article": "docs/support/exotel-vocallabs-sip-trunk.md",
        "pattern": "Superflow B2B API + createSIPCall / websockets",
        "dashboard": "https://docs.vocallabs.ai/vocallabs",
        "docs": "https://api.superflow.run/b2b/",
        "notes": "API-first; classic vSIP only if vendor provides SIP target",
    },
    {
        "name": "Rapida AI",
        "article": "docs/support/exotel-rapida-ai-sip-trunk.md",
        "pattern": "Native Exotel stream OR SIP to sip-01.in.rapida.ai",
        "dashboard": "https://www.rapida.ai/",
        "docs": "https://doc.rapida.ai/integrations/telephony/exotel",
        "notes": "Path A: webhook/stream; Path B: vSIP + destination-uris",
    },
    {
        "name": "NLPearl.AI",
        "article": "docs/support/exotel-nlpearl-sip-trunk.md",
        "pattern": "Custom VoIP (SIP domain inbound + SIP trunk URL outbound)",
        "dashboard": "https://platform.nlpearl.ai/",
        "docs": "https://developers.nlpearl.ai/pages/custom_voip",
        "notes": "Option B: Exotel vSIP behind NLPearl; inbound uses destination-uris to SIP Domain; outbound uses trunk URL",
    },
]

MATRIX_DIMS = [
    ("Exotel vSIP trunk API", "vSIP_API"),
    ("Outbound SIP (AI→Exotel→PSTN)", "out_sip"),
    ("Inbound SIP (PSTN→Exotel→partner)", "in_sip"),
    ("Digest /credentials match", "digest"),
    ("whitelisted-ips (static /32)", "acl"),
    ("destination-uris + Connect sip:<trunk_sid>", "dest_connect"),
    ("Webhook / streaming (non-SIP trunk)", "webhook_stream"),
    ("Programmatic API outbound", "api_out"),
]

MATRIX_ROWS = [
    {
        "provider": "ElevenLabs",
        "vSIP_API": "Y",
        "out_sip": "Y",
        "in_sip": "Y",
        "digest": "Y",
        "acl": "Optional",
        "dest_connect": "Y (inbound)",
        "webhook_stream": "N",
        "api_out": "Y (ConvAI)",
    },
    {
        "provider": "LiveKit Cloud",
        "vSIP_API": "Y",
        "out_sip": "Y",
        "in_sip": "Y",
        "digest": "Y",
        "acl": "Optional",
        "dest_connect": "Y (inbound)",
        "webhook_stream": "N",
        "api_out": "Y (SDK)",
    },
    {
        "provider": "Retell AI",
        "vSIP_API": "Y",
        "out_sip": "Y",
        "in_sip": "Y",
        "digest": "Y",
        "acl": "Optional",
        "dest_connect": "Y",
        "webhook_stream": "N",
        "api_out": "Y",
    },
    {
        "provider": "Bolna",
        "vSIP_API": "Y",
        "out_sip": "Y",
        "in_sip": "Y",
        "digest": "Y",
        "acl": "Optional",
        "dest_connect": "Y",
        "webhook_stream": "N (BYOT path)",
        "api_out": "Y",
    },
    {
        "provider": "Pipecat + Daily",
        "vSIP_API": "Y (PSTN leg)",
        "out_sip": "Partial",
        "in_sip": "Partial",
        "digest": "Y",
        "acl": "Optional",
        "dest_connect": "If fixed SIP design",
        "webhook_stream": "Y (typical)",
        "api_out": "Y",
    },
    {
        "provider": "Ultravox",
        "vSIP_API": "Y (SIP path)",
        "out_sip": "Y",
        "in_sip": "Y",
        "digest": "Y",
        "acl": "Optional",
        "dest_connect": "Y",
        "webhook_stream": "Alt (exotel medium)",
        "api_out": "Y",
    },
    {
        "provider": "Vapi",
        "vSIP_API": "Y",
        "out_sip": "Y",
        "in_sip": "Y",
        "digest": "Y",
        "acl": "Recommended (2× Vapi SBC /32)",
        "dest_connect": "Y",
        "webhook_stream": "N",
        "api_out": "Y",
    },
    {
        "provider": "Smallest AI (Atoms)",
        "vSIP_API": "Y",
        "out_sip": "Y",
        "in_sip": "Y",
        "digest": "Y",
        "acl": "Optional",
        "dest_connect": "Y",
        "webhook_stream": "N",
        "api_out": "Y",
    },
    {
        "provider": "Vocallabs",
        "vSIP_API": "If SIP design",
        "out_sip": "Partial",
        "in_sip": "Partial",
        "digest": "If SIP",
        "acl": "If static egress known",
        "dest_connect": "If SIP URI provided",
        "webhook_stream": "Y (typical)",
        "api_out": "Y",
    },
    {
        "provider": "Rapida AI",
        "vSIP_API": "Path B only",
        "out_sip": "Path B",
        "in_sip": "Path B",
        "digest": "Path B",
        "acl": "Optional",
        "dest_connect": "Path B",
        "webhook_stream": "Path A (native Exotel)",
        "api_out": "Y",
    },
    {
        "provider": "NLPearl.AI",
        "vSIP_API": "Y",
        "out_sip": "Y",
        "in_sip": "Y",
        "digest": "Y",
        "acl": "Avoid (unless static /32 needed)",
        "dest_connect": "Y (inbound)",
        "webhook_stream": "N",
        "api_out": "Y",
    },
]

EXOTEL_COMMON = [
    ("EX-001", "Prerequisites", "Exotel account: KYC done, vSIP enabled, Exophone in E.164", "Account can create trunk; DID visible", "All vSIP paths"),
    ("EX-002", "API auth", "Call trunk API with Basic auth (API Key : API Token)", "200 responses; no 401", "All vSIP API use"),
    ("EX-003", "Create trunk", "POST /v2/accounts/{sid}/trunks", "trunk_sid returned", "All vSIP"),
    ("EX-004", "Map DID", "POST .../phone-numbers with E.164", "Phone mapped to trunk", "All vSIP"),
    ("EX-005", "SIP digest", "POST .../credentials; match partner", "401 absent on good call", "Digest-based"),
    ("EX-006", "Whitelist static IP", "POST .../whitelisted-ips with mask 32 per IP", "Partner signalling allowed", "When partner has static egress"),
    ("EX-007", "Inbound destination", "POST .../destination-uris toward partner SIP host", "INVITE reaches partner", "Inbound PSTN flows"),
    ("EX-008", "Flow Connect", "Dial whom = sip:<trunk_sid> only", "Call routes per flow", "Inbound using Connect"),
    ("EX-009", "Edge IP:port", "Use Exotel-assigned edge; may be in.voip.exotel.com:5070 (TCP) or :443 (TLS)", "Partner can complete SIP to Exotel", "Outbound from partner"),
    ("EX-010", "Negative: bad digest", "Mismatch password between Exotel and partner", "401/auth failure", "Regression"),
]


def build_workbook() -> Workbook:
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Instructions"
    ws0["A1"] = "Exotel vSIP + Voice AI — test matrix"
    ws0["A1"].font = Font(bold=True, size=14)
    lines = [
        f"Generated: {date.isoformat(date.today())}",
        "Repo: Voice AI Ecosystem — docs/support/",
        "",
        "How to use:",
        "1) Start with Provider_Catalog and Coverage_Matrix.",
        "2) Run Exotel_Common_Tests for any integration using vSIP trunk APIs.",
        "3) Execute Detailed_Test_Cases for your provider; capture evidence links.",
        "4) Use Regression_Checklist for release sign-off.",
        "",
        "Regenerate: .venv-xlsx/bin/python scripts/generate-voice-ai-test-matrix.py",
    ]
    for i, line in enumerate(lines, start=2):
        ws0[f"A{i}"] = line
    ws0.column_dimensions["A"].width = 92

    ws1 = wb.create_sheet("Provider_Catalog")
    ws1.append(["Provider", "Repo_support_article", "Integration_pattern", "Dashboard_URL", "Official_docs_URL", "Notes"])
    style_header(ws1)
    for p in PROVIDERS:
        ws1.append([p["name"], p["article"], p["pattern"], p["dashboard"], p["docs"], p["notes"]])
    autosize_columns(ws1)

    ws2 = wb.create_sheet("Coverage_Matrix")
    ws2.append(["Provider"] + [d[0] for d in MATRIX_DIMS])
    style_header(ws2)
    keys = ["provider"] + [d[1] for d in MATRIX_DIMS]
    for row in MATRIX_ROWS:
        ws2.append([row[k] for k in keys])
    autosize_columns(ws2)

    ws3 = wb.create_sheet("Exotel_Common_Tests")
    ws3.append(["Test_ID", "Category", "Test_case", "Expected", "Applies_to"])
    style_header(ws3)
    for r in EXOTEL_COMMON:
        ws3.append(list(r))
    autosize_columns(ws3)

    ws4 = wb.create_sheet("Detailed_Test_Cases")
    ws4.append(
        [
            "Test_ID",
            "Provider",
            "Layer",
            "Direction",
            "Test_title",
            "Preconditions",
            "Steps_summary",
            "Expected_result",
            "Pass_Fail",
            "Run_date",
            "Tester",
            "Evidence_link",
            "Notes",
        ]
    )
    style_header(ws4)

    tpl = [
        ("{pfx}-01", "Prereq", "Both", "Account + product gates", "KYC, vSIP, DID, partner account", "Verify in consoles", "Ready to configure"),
        ("{pfx}-02", "Exotel", "N/A", "Create + map DID", "API credentials", "Snippets: trunk + phone-numbers", "trunk_sid + DID on trunk"),
        ("{pfx}-03", "Exotel", "Outbound", "POST credentials", "Digest chosen", "Match partner-side auth", "401 absent on good call"),
        ("{pfx}-04", "Partner", "Outbound", "Configure termination to Exotel edge", "Edge per Exotel (IP:port or in.voip.exotel.com)", "Per provider doc", "SIP reaches Exotel"),
        ("{pfx}-05", "E2E", "Outbound", "PSTN call completes", "Steps 02–04 done", "Place call to mobile", "Ring/answer; audio path"),
        ("{pfx}-06", "Partner", "Inbound", "Set inbound target on partner", "If inbound in scope", "Destination/origination per doc", "Partner accepts from Exotel"),
        ("{pfx}-07", "Exotel", "Inbound", "destination-uris + Flow", "Trunk ready", "POST destination-uris; Connect sip:<trunk_sid>", "PSTN→Exotel→partner"),
        ("{pfx}-08", "E2E", "Inbound", "Caller reaches agent", "Inbound wired", "Dial Exotel DID", "Agent answers; 2-way audio"),
        ("{pfx}-09", "Security", "Both", "ACL if required", "Static IPs known", "whitelisted-ips /32 each", "No spurious 403"),
        ("{pfx}-10", "Regression", "Both", "Codec / media sanity", "Call up", "Listen + interrupt", "Clear audio; acceptable latency"),
    ]

    prefixes = {
        "ElevenLabs": "EL",
        "LiveKit Cloud": "LK",
        "Retell AI": "RT",
        "Bolna": "BL",
        "Pipecat + Daily": "PC",
        "Ultravox": "UV",
        "Vapi": "VP",
        "Smallest AI (Atoms)": "SM",
        "Vocallabs": "VC",
        "Rapida AI": "RP",
        "NLPearl.AI": "NP",
    }

    for p in PROVIDERS:
        name = p["name"]
        pfx = prefixes.get(name, name[:2].upper())
        for tid, layer, direction, title, pre, steps, exp in tpl:
            notes = ""
            if name == "Pipecat + Daily" and "destination-uris" in steps:
                notes = "Often bridge to Daily sip_uri instead of static destination-uris-only"
            if name == "Vocallabs" and layer == "Partner":
                notes = "Confirm SIP vs API-first path with Vocallabs"
            if name == "Rapida AI" and "Exotel" in layer:
                notes = "Path A may skip trunk APIs; use Rapida Exotel doc"
            if name == "Vapi" and "ACL" in title:
                notes = "Whitelist both Vapi SBC IPs"
            ws4.append([tid.format(pfx=pfx), name, layer, direction, title, pre, steps, exp, "", "", "", "", notes])

    autosize_columns(ws4)

    ws5 = wb.create_sheet("Regression_Checklist")
    ws5.append(["Provider", "Smoke_outbound_OK", "Smoke_inbound_OK", "Docs_reviewed", "Notes_this_run", "Signoff", "Date"])
    style_header(ws5)
    for p in PROVIDERS:
        ws5.append([p["name"], "", "", "", "", "", ""])
    autosize_columns(ws5)

    ws6 = wb.create_sheet("Other_Integrations")
    ws6.append(["Area", "Doc_in_repo", "Test_focus"])
    style_header(ws6)
    ws6.append(["WebRTC / Exotel Voice APIs (browser SDK)", "docs/integrations/webrtc-application-setup.md", "initWebrtc, DoRegister, outbound connect, Agent-Stream bridge"])
    ws6.append(["Scripts — Exotel + ElevenLabs", "scripts/exotel-elevenlabs/README.md", "Run 01–04 scripts; .env not committed"])
    ws6.append(["LiveKit outbound sample", "Livekit/livekit-outbound-caller-agent/OUTBOUND-EXOTEL-NOTES.md", "403, E.164, no-audio patterns"])
    autosize_columns(ws6)

    return wb


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()

